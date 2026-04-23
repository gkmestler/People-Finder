"""Build formatted Excel spreadsheet from enrichment results."""

from __future__ import annotations

import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
GREEN_FILL = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
PINK_FILL = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
GRAY_FILL = PatternFill(start_color="e2e3e5", end_color="e2e3e5", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

HEADERS = ["Company", "First Name", "Last Name", "Title", "Email", "Email Status", "LinkedIn URL"]
HEADERS_WITH_PHONE = ["Company", "First Name", "Last Name", "Title", "Phone", "Email", "Email Status", "LinkedIn URL"]
HEADERS_SEARCH_ONLY = ["Company", "First Name", "Last Name", "Title", "LinkedIn URL"]
COL_WIDTHS = [35, 15, 20, 45, 40, 15, 55]
COL_WIDTHS_WITH_PHONE = [35, 15, 20, 45, 20, 40, 15, 55]
COL_WIDTHS_SEARCH_ONLY = [35, 15, 20, 45, 55]


def build_spreadsheet(contacts: list[dict], no_results: list[str], output_path: str | None = None, include_phone: bool = False, search_only: bool = False) -> str | io.BytesIO:
    """Build a formatted Excel spreadsheet from enrichment results.

    Args:
        contacts: list of enriched contact dicts
        no_results: list of company names with no Apollo results
        output_path: file path to save the .xlsx, or None to return a BytesIO object
        search_only: if True, only include name/title/company/LinkedIn (no email/phone)

    Returns:
        output_path if saving to disk, or BytesIO object if output_path is None
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    if search_only:
        headers = HEADERS_SEARCH_ONLY
        col_widths = COL_WIDTHS_SEARCH_ONLY
    elif include_phone:
        headers = HEADERS_WITH_PHONE
        col_widths = COL_WIDTHS_WITH_PHONE
    else:
        headers = HEADERS
        col_widths = COL_WIDTHS

    # --- Headers ---
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 28

    # --- Sort contacts by company name ---
    contacts_sorted = sorted(contacts, key=lambda c: (c.get("organization_name") or "", c.get("last_name") or ""))

    num_cols = len(headers)

    # --- Data rows ---
    for i, contact in enumerate(contacts_sorted, 2):
        if search_only:
            row_data = [
                contact.get("organization_name", ""),
                contact.get("first_name", ""),
                contact.get("last_name", ""),
                contact.get("title", ""),
                contact.get("linkedin_url", ""),
            ]
            row_fill = GREEN_FILL
        else:
            row_data = [
                contact.get("organization_name", ""),
                contact.get("first_name", ""),
                contact.get("last_name", ""),
                contact.get("title", ""),
            ]
            if include_phone:
                row_data.append(contact.get("phone_number") or "")
            row_data.extend([
                contact.get("email") or "",
                contact.get("email_status", ""),
                contact.get("linkedin_url", ""),
            ])

            email = contact.get("email")
            status = contact.get("email_status", "")

            if email and status in ("verified", "extrapolated"):
                row_fill = GREEN_FILL
            elif status == "error":
                row_fill = PINK_FILL
            else:
                row_fill = YELLOW_FILL

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = BODY_FONT
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    # --- No results rows ---
    start_row = len(contacts_sorted) + 2
    for j, company in enumerate(no_results):
        row = start_row + j
        ws.cell(row=row, column=1, value=company).font = BODY_FONT
        ws.cell(row=row, column=4, value="NO RESULTS IN APOLLO").font = Font(name="Arial", size=10, italic=True)
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PINK_FILL
            cell.border = THIN_BORDER

    # --- Legend sheet ---
    legend = wb.create_sheet("Legend")
    legend_data = [
        ("Color", "Meaning"),
        ("Green", "Enriched with email found (verified or extrapolated)"),
        ("Yellow", "Enriched but email unavailable"),
        ("Pink", "No results found in Apollo / enrichment error"),
    ]
    for r, (label, meaning) in enumerate(legend_data, 1):
        legend.cell(row=r, column=1, value=label).font = Font(bold=(r == 1), name="Arial", size=10)
        legend.cell(row=r, column=2, value=meaning).font = Font(bold=(r == 1), name="Arial", size=10)
    legend.column_dimensions["A"].width = 12
    legend.column_dimensions["B"].width = 55

    # Color the legend cells
    legend.cell(row=2, column=1).fill = GREEN_FILL
    legend.cell(row=3, column=1).fill = YELLOW_FILL
    legend.cell(row=4, column=1).fill = PINK_FILL

    # --- Stats sheet ---
    stats_sheet = wb.create_sheet("Stats")
    stats_data = [
        ("Metric", "Value"),
        ("Total Contacts", len(contacts_sorted)),
        ("With Email", sum(1 for c in contacts_sorted if c.get("email"))),
        ("Verified Emails", sum(1 for c in contacts_sorted if c.get("email_status") == "verified")),
        ("Extrapolated Emails", sum(1 for c in contacts_sorted if c.get("email_status") == "extrapolated")),
        ("No Email", sum(1 for c in contacts_sorted if not c.get("email"))),
        ("Companies with Results", len(set(c.get("organization_name", "") for c in contacts_sorted))),
        ("Companies with No Results", len(no_results)),
    ]
    for r, (label, val) in enumerate(stats_data, 1):
        stats_sheet.cell(row=r, column=1, value=label).font = Font(bold=(r == 1), name="Arial", size=10)
        stats_sheet.cell(row=r, column=2, value=val).font = Font(bold=(r == 1), name="Arial", size=10)
    stats_sheet.column_dimensions["A"].width = 28
    stats_sheet.column_dimensions["B"].width = 15

    # --- Freeze panes ---
    ws.freeze_panes = "A2"

    # --- Auto-filter ---
    last_col = chr(ord('A') + num_cols - 1)
    ws.auto_filter.ref = f"A1:{last_col}{len(contacts_sorted) + 1}"

    if output_path is None:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path
